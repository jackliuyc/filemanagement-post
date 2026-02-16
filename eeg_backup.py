import sys
import json
import re
import os
import shutil
import pandas as pd

import ulid

from datetime import datetime

from openpyxl import load_workbook
from zipfile import ZipFile

from PyQt5.QtCore import pyqtSignal, QDate, Qt
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QVBoxLayout,
    QHBoxLayout,
    QFormLayout,
    QWidget,
    QLabel,
    QLineEdit,
    QComboBox,
    QPushButton,
    QDateEdit,
    QSpinBox,
    QMessageBox,
    QScrollArea,
    QFrame,
    QFileDialog,
    QTabWidget,
    QSizePolicy,
    QAction,
    QProgressBar,
    QDialog,
)


class SessionInfoForm(QWidget):

    # signal to main window when session info is confirmed
    confirm_session_info_signal = pyqtSignal()

    def __init__(self, data_model=None, parent=None):
        super().__init__(parent)

        # Init data model
        self.data_model = data_model

        # Inputs and indicator labels
        self.inputs = {}
        self.indicators = {}

        # Layout setup
        self.layout = QVBoxLayout(self)

        # Preset combo box
        self.preset_combo = QComboBox()
        self.preset_combo.addItems(self.data_model.config_dict.keys())
        self.preset_combo.currentTextChanged.connect(self.load_preset)
        self.layout.addWidget(QLabel("Select Study Preset:"))
        self.layout.addWidget(self.preset_combo)

        # Scroll area for input fields
        self.scroll_area = QScrollArea()
        self.scroll_area.setStyleSheet("background-color: white")
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.scroll_area.setMinimumHeight(400)
        self.scroll_content = QWidget()
        self.scroll_area.setWidget(self.scroll_content)
        self.layout.addWidget(self.scroll_area)

        # Lock filename button
        self.confirm_session_button = QPushButton("Confirm Session Information")
        self.confirm_session_button.clicked.connect(
            self.confirm_session_info_signal.emit
        )
        self.confirm_session_button.setEnabled(
            False
        )  # not enabled until fields are validated
        self.confirm_session_button.setStyleSheet(
            """
            QPushButton:disabled {
                background-color: #A0A0A0;
            }
        """
        )
        self.layout.addWidget(self.confirm_session_button)

        # Load first preset
        self.reset_session_form()

    def get_current_study(self):
        """Get current study preset from combobox text. Could alternatively get from data model?"""
        return self.preset_combo.currentText()

    def reset_session_form(self):
        """Reset to default form and current study"""
        self.preset_combo.setCurrentIndex(0)
        self.load_preset(self.get_current_study())
        self.update_session_info()  # should be blank

    def update_session_info(self):
        """Update data model with currently entered session information"""
        cur_session_info = self.data_model.session_info
        for key in cur_session_info:
            value = self.get_input_value(key)
            self.data_model.session_info[key] = value

    def load_preset(self, preset):
        """Clear and reset all form elements, clear input and indicators, then load UI elements of preset from configuration dict"""

        # Clear all elements and indicators in form
        if self.scroll_content.layout() is not None:
            # Clear the existing layouts
            for i in reversed(range(self.scroll_content.layout().count())):
                layout = self.scroll_content.layout().itemAt(i).layout()
                if layout is not None:
                    for j in reversed(range(layout.rowCount())):
                        layout.removeRow(j)
            # Remove the column layouts
            while self.scroll_content.layout().count() > 0:
                item = self.scroll_content.layout().takeAt(0)
                if item.layout():
                    item.layout().setParent(None)
        self.inputs.clear()
        self.indicators.clear()

        # Load preset from data model
        preset = self.data_model.config_dict[preset]

        # Create a new widget for the scroll area content
        self.scroll_content = QWidget()

        # Create a horizontal layout for two columns
        columns_layout = QHBoxLayout(self.scroll_content)
        left_form = QFormLayout()
        right_form = QFormLayout()
        columns_layout.addLayout(left_form)
        columns_layout.addLayout(right_form)

        # Set the new widget as the scroll area's widget
        self.scroll_area.setWidget(self.scroll_content)

        total_fields = len(preset)
        fields_per_column = (total_fields + 1) // 2  # Round up division
        field_count = 0

        for field_name, field in preset.items():
            if field["type"] == "hidden":
                continue  # Skip hidden fields

            widget = self.create_widget(field)
            row_layout = QVBoxLayout()  # Changed to QVBoxLayout

            # Add error label
            error_label = QLabel()
            error_label.setStyleSheet("color: red; font-size: 12px;")
            error_label.setVisible(False)
            row_layout.addWidget(error_label)

            widget_row = QHBoxLayout()
            widget_row.addWidget(widget)
            widget_row.addStretch()  # Add stretch to push widget to the left

            if field.get("editable", True):
                indicator = QLabel("❌")  # Red X
                indicator.setStyleSheet("color: red; font-size: 16px;")
                self.indicators[field_name] = indicator
                widget_row.addWidget(indicator)

            row_layout.addLayout(widget_row)

            if field_count < fields_per_column:
                left_form.addRow(field["label"], row_layout)
            else:
                right_form.addRow(field["label"], row_layout)

            self.inputs[field_name] = {"widget": widget, "error_label": error_label}
            field_count += 1

        # Adjust the scroll content widget's layout
        self.scroll_content.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.scroll_content.setLayout(columns_layout)

        # Validate all fields after loading
        self.validate_all_fields()

    def create_widget(self, field):
        """Create widget and connect relevant signals to validate_all_fields."""
        if field["type"] == "text":
            widget = QLineEdit()
            widget.setMinimumHeight(30)
            if "default" in field:
                widget.setText(field["default"])
            if "editable" in field and not field["editable"]:
                widget.setReadOnly(True)
            else:
                widget.textChanged.connect(
                    self.validate_all_fields
                )  # Connect to validate_all_fields
        elif field["type"] == "combo":
            widget = QComboBox()
            widget.setMinimumHeight(30)
            widget.addItems(field["options"])
            if field.get("editable", True):
                widget.currentTextChanged.connect(
                    self.validate_all_fields
                )  # Connect to validate_all_fields
            else:
                widget.setEnabled(False)
        elif field["type"] == "date":
            widget = QDateEdit()
            widget.setMinimumHeight(30)
            widget.setCalendarPopup(True)
            widget.setDate(QDate.currentDate())
            if field.get("editable", True):
                widget.dateChanged.connect(
                    self.validate_all_fields
                )  # Connect to validate_all_fields
            else:
                widget.setReadOnly(True)
        elif field["type"] == "spinbox":
            widget = QSpinBox()
            widget.setMinimumHeight(30)
            widget.setMinimum(0)
            widget.setMaximum(99999)
            if field.get("editable", True):
                widget.valueChanged.connect(
                    self.validate_all_fields
                )  # Connect to validate_all_fields
            else:
                widget.setReadOnly(True)
        elif field["type"] == "hidden":
            widget = QLineEdit()
            widget.setVisible(False)

        return widget

    def update_indicator(self, field_name, is_valid):
        """Update form to reflect indicators"""
        if field_name in self.indicators:
            self.indicators[field_name].setText("✅" if is_valid else "❌")
            self.indicators[field_name].setStyleSheet(
                "color: green;" if is_valid else "color: red;"
            )

    def get_input_value(self, name):
        """Get value from given element"""
        widget = self.inputs[name]["widget"]
        if isinstance(widget, QLineEdit):
            return widget.text()
        elif isinstance(widget, QComboBox):
            return widget.currentText()
        elif isinstance(widget, QDateEdit):
            return widget.date().toString("MM-dd-yyyy")
        elif isinstance(widget, QSpinBox):
            return str(widget.value())
        return ""

    def validate_all_fields(self):
        """Validate all fields and update the overall UI state based on validation results."""
        preset = self.data_model.config_dict[self.get_current_study()]
        all_valid = True

        # Loop over all fields in the preset
        for field_name, field in preset.items():
            if field["type"] == "hidden":
                continue  # Skip hidden fields

            # Get the current value of the field
            value = self.get_input_value(field_name)

            # Get info from config dict for validation
            widget = self.inputs[field_name]["widget"]
            error_label = self.inputs[field_name]["error_label"]

            # Check if the field's value matches the regex validation
            if "validation" in field:
                is_valid = re.match(field["validation"], value) is not None
                if is_valid:
                    widget.setStyleSheet("border: 1px solid green;")
                    self.update_indicator(field_name, True)
                    error_label.setVisible(False)
                else:
                    widget.setStyleSheet("border: 1px solid red;")
                    self.update_indicator(field_name, False)
                    error_label.setText(
                        field.get("error_message", "This field is required")
                    )
                    error_label.setVisible(True)
                    all_valid = False  # Mark form as invalid if any field is invalid

        # Enable or disable the confirm button based on the overall validation result
        self.confirm_session_button.setEnabled(all_valid)


class FileInputForm(QWidget):

    confirm_file_info_signal = pyqtSignal()

    def __init__(self, data_model=None, parent=None):
        super().__init__(parent)

        # Set data model
        self.data_model = data_model

        # List to hold all sections
        self.sections = []

        # Main layout
        self.layout = QVBoxLayout(self)

        # Notes file input
        self.notes_button = QPushButton("Load Session Notes (.txt, .rtf)")
        self.notes_button.clicked.connect(self.upload_notes_file)
        self.notes_label = QLabel("No file selected")
        self.layout.addWidget(self.notes_button)
        self.layout.addWidget(self.notes_label)

        # Net placement photos
        self.photos_button = QPushButton(
            "Load Net Placement Photos (.png, .jpg, .jpeg)"
        )
        self.photos_button.clicked.connect(self.upload_photos)
        self.photos_label = QLabel("No photos selected")
        self.layout.addWidget(self.photos_button)
        self.layout.addWidget(self.photos_label)

        # Container for sections
        self.scroll_area = QScrollArea()
        self.scroll_widget = QWidget()
        self.scroll_layout = QVBoxLayout()
        self.scroll_widget.setLayout(self.scroll_layout)
        self.scroll_area.setWidget(self.scroll_widget)
        self.scroll_area.setWidgetResizable(True)
        self.layout.addWidget(self.scroll_area)

        # Add paradigm button
        self.add_button = QPushButton("Add Additional Paradigms")
        self.add_button.clicked.connect(self.add_section)
        self.add_button.setEnabled(False)
        self.layout.addWidget(self.add_button)

        # Add confirm session info button
        self.confirm_file_button = QPushButton("Confirm and Upload")
        self.confirm_file_button.clicked.connect(self.confirm_file_info_signal)
        self.confirm_file_button.setEnabled(False)
        self.layout.addWidget(self.confirm_file_button)

        # Add initial section
        self.add_section()  # Initialize with one section

    def add_section(self):
        """Add section for EEG paradigm. Contains paradigm combobox and buttons for MFF file selection"""
        # Layout
        form_layout = QFormLayout()

        # Paradigm selection combo box
        paradigm_combo = QComboBox()
        paradigm_combo.wheelEvent = lambda event: None
        paradigm_combo.addItems(self.data_model.get_list_of_current_paradigms())
        paradigm_combo.currentIndexChanged.connect(self.check_form_completion)
        form_layout.addRow(
            QLabel(f"Paradigm {len(self.sections) + 1}:"), paradigm_combo
        )
        
        # New combo box row for audio source for each paradigm
        audio_combo = QComboBox()
        audio_combo.wheelEvent = lambda event: None
        audio_combo.addItems(["none", "headphones", "speakers"])
        audio_combo.currentIndexChanged.connect(self.check_form_completion)
        form_layout.addRow(
            QLabel("Audio source:"), audio_combo
        )

        # MFF file button
        mff_button = QPushButton("Upload .MFF file")
        mff_label = QLabel("No file selected")
        mff_button.clicked.connect(
            lambda _, label=mff_label, combo=paradigm_combo: self.upload_mff(
                label, combo
            )
        )
        form_layout.addRow(mff_button, mff_label)

        # Sections for each paradigm using QWidgets
        section_widget = QWidget()
        section_widget.setLayout(form_layout)
        self.sections.append(
            {
                "paradigm_combo": paradigm_combo,
                "audio_combo": audio_combo,
                "mff_label": mff_label,
                "widget": section_widget,
            }
        )
        self.scroll_layout.addWidget(section_widget)

        # Add a divider
        divider = QFrame()
        divider.setFrameShape(QFrame.HLine)
        divider.setFrameShadow(QFrame.Sunken)
        self.scroll_layout.addWidget(divider)

        # Make sure buttons are updated
        self.check_form_completion()

    def check_form_completion(self):
        """Enable or disable buttons depending on if file selection is complete for each section."""
        # Check if notes present
        if self.notes_label.text() == "No file selected":
            self.add_button.setEnabled(False)
            self.confirm_file_button.setEnabled(False)
            return
        # Check if all sections are complete (selected paradigm + loaded mff)
        all_sections_complete = all(
            section["paradigm_combo"].currentIndex() != 0
            and section["mff_label"].text() != "No file selected"
            for section in self.sections
        )
        self.add_button.setEnabled(all_sections_complete)
        self.confirm_file_button.setEnabled(all_sections_complete)

    def upload_mff(self, mff_label, paradigm_combo):
        """File dialog for selecting MFF file. Checks that selected folder is .mff and if file name matches chosen paradigm."""
        options = QFileDialog.Options()
        default_folder = self.data_model.filepath_dict["usb_input_dir"]
        if not os.path.exists(default_folder):
            default_folder = ""
        folder = QFileDialog.getExistingDirectory(
            self, "Select .MFF file", default_folder, options=options
        )

        if folder:
            # check if the folder name ends with .mff
            folder_name = os.path.basename(folder)
            if folder_name.endswith(".mff"):

                # check if folder name contains paradigm
                paradigm_name = paradigm_combo.currentText()
                if not paradigm_name.lower() in folder_name.lower():
                    QMessageBox.warning(
                        self,
                        "WARNING",
                        f'The .mff file you loaded does not match the EEG paradigm: "{paradigm_name}"\n\nDouble check that you have selected the correct .mff file!',
                    )

                # set label regardless (only warn do not force user to have paradigm in file name in case of typos)
                mff_label.setText(folder)

            else:
                mff_label.setText("No file selected")
                QMessageBox.warning(
                    self, "WARNING", "The selected file is not a valid .mff file!"
                )
        self.check_form_completion()  # Check validity and update buttons

    def upload_notes_file(self):
        """Open file dialog for user to select a notes file"""
        options = QFileDialog.Options()
        default_folder = self.data_model.filepath_dict["usb_input_dir"]
        if not os.path.exists(default_folder):
            default_folder = ""
        filename, _ = QFileDialog.getOpenFileName(
            self,
            "Select a file",
            default_folder,
            "Text Files (*.txt;*.rtf);;All Files (*)",
            options=options,
        )
        if filename:
            self.data_model.notes_file = filename
            self.notes_label.setText(filename)
            self.check_form_completion()

    def upload_photos(self):
        """Open file dialog for user to select net placement photos"""
        options = QFileDialog.Options()
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "Select Images",
            "",
            "Images (*.png *.jpg *.jpeg);;All Files (*)",
            options=options,
        )
        if files:
            self.data_model.net_placement_photos = files
            self.photos_label.setText(f"{len(files)} images selected")
        else:
            self.photos_label.setText("No photos selected")
            self.check_form_completion()

    def reset_file_form(self):
        """Clear all elements and data in file info form form"""

        # Clear notes and photos data
        self.data_model.notes_file = None  # Reset notes file
        self.notes_label.setText("No file selected")  # Reset label for notes

        self.data_model.net_placement_photos = []  # Reset photos data
        self.photos_label.setText("No photos selected")  # Reset label for photos

        # Remove all widgets and dividers from the scroll layout
        while self.scroll_layout.count():
            item = self.scroll_layout.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()

        # Clear and re-init first section
        self.sections = []
        self.add_section()

        # Reset buttons
        self.add_button.setEnabled(False)
        self.confirm_file_button.setEnabled(False)

    def update_file_info(self):
        """Update data_model user input fields."""
        # Clear the existing eeg_file_info
        self.data_model.eeg_file_info = []

        # Loop over each section
        for section in self.sections:

            # Dictionary of file info
            paradigm = section["paradigm_combo"].currentText()
            audio_source = section["audio_combo"].currentText()
            mff_file = section["mff_label"].text()
            file_info = {
                "paradigm": paradigm,
                "audio_source": audio_source,
                "mff_file": mff_file if mff_file != "No file selected" else None,
            }
            self.data_model.eeg_file_info.append(file_info)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        # Initialize data model
        self.data_model = DataModel()

        # Set stylesheet
        self.setStyleSheet(
            """
            QMainWindow {
                background-color: #E6F3FF;
            }
            QLabel {
                font-size: 16px;
                color: #1A3A54;
            }
            QComboBox, QLineEdit, QDateEdit, QSpinBox {
                font-size: 16px;
                padding: 5px;
                border: 1px solid #4A90E2;
                border-radius: 4px;
                background-color: #FFFFFF;
                color: #1A3A54;
                width: 150;
            }
            QPushButton {
                font-size: 16px;
                padding: 10px;
                background-color: #4A90E2;
                color: white;
                border: none;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #3A7BC8;
            }
            QPushButton:disabled {
                background-color: #B0B0B0;  
                color: #7F7F7F;      
            }
            QCheckBox {
                font-size: 16px;
                color: #1A3A54;
            }
            QTabBar::tab {
                font-size: 16px;
                min-width: 180;  
                padding: 5px; 
                border: 1px solid #D1D9E6;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
            }
            QTabBar::tab:selected {
                background-color: #FFFFFF;
                border-bottom-color: #FFFFFF;
            }
        """
        )
        self.setWindowTitle("Main Application")
        self.setGeometry(100, 100, 800, 600)

        # Create the menu
        self.init_menu()

        # Tab widget
        self.tab_widget = QTabWidget()
        self.setCentralWidget(self.tab_widget)

        # Init session info tab
        self.session_info_tab = SessionInfoForm(self.data_model)
        self.tab_widget.addTab(self.session_info_tab, "Session Information")

        # Slot for session info confirm signal
        self.session_info_tab.confirm_session_info_signal.connect(
            self.validate_session_and_swap_tabs
        )

        # Init file upload tab
        self.file_upload_tab = FileInputForm(self.data_model)
        self.tab_widget.addTab(self.file_upload_tab, "File Upload")
        self.tab_widget.setTabEnabled(0, True)  # Enable first tab
        self.tab_widget.setTabEnabled(1, False)  # Disable second tab initially
        self.tab_widget.setCurrentIndex(0)

        # Slot for file info confirm signal
        self.file_upload_tab.confirm_file_info_signal.connect(self.process_files)

    def init_menu(self):
        """Create menu bar with reset form and select output items"""
        menu_bar = self.menuBar()
        file_menu = menu_bar.addMenu("File")

        # Reset form
        reset_action = QAction("Reset Form", self)
        reset_action.triggered.connect(self.reset_app)
        file_menu.addAction(reset_action)

        # Quit application
        quit_action = QAction("Quit", self)
        quit_action.triggered.connect(QApplication.quit)
        file_menu.addAction(quit_action)

    def reset_app(self):
        """Reset all fields and data model"""
        # Reset the session info tab and file tab
        self.session_info_tab.reset_session_form()
        self.file_upload_tab.reset_file_form()

        # Disable the second tab
        self.tab_widget.setTabEnabled(1, False)

        # Swap to first tab
        self.tab_widget.setTabEnabled(0, True)
        self.tab_widget.setCurrentIndex(0)

        # Clear data amodel
        self.data_model.clear_data()

    def validate_session_and_swap_tabs(self):
        """When session confirm button is clicked: double check validity, update model, and swap to second tab"""
        all_valid = all(
            self.session_info_tab.indicators[field].text() == "✅"
            for field in self.session_info_tab.indicators
        )
        if not all_valid:
            QMessageBox.warning(
                self,
                "WARNING",
                "Check that the fields are valid. If you see this something is really really wrong.",
            )
            return

        # Update data model with session information
        self.session_info_tab.update_session_info()

        # Check if session info already exists in deid log
        if self.data_model.check_if_session_info_already_exists():
            QMessageBox.warning(
                self,
                "WARNING",
                "The session information matches an existing entry in the DeID log! Cannot proceed with duplicate Check that you entered everything correctly.",
            )
            return

        # Swap to second tab
        self.file_upload_tab.reset_file_form()
        self.tab_widget.setTabEnabled(1, True)
        self.tab_widget.setCurrentIndex(1)
        # Disable first tab (have to reset whole session)
        self.tab_widget.setTabEnabled(0, False)

    def ask_user_for_file_confirmation(self):
        # join paradigms and file names
        paradigm_names_string = "\n ".join(
            [
                os.path.basename(section["mff_label"].text())
                for section in self.file_upload_tab.sections
            ]
        )
        # ask user for confirmation
        message = f"Confirm that you want to upload the following {len(self.file_upload_tab.sections)} file(s)?\n\n{paradigm_names_string}"
        reply = QMessageBox.question(
            self,
            "Confirmation",
            message,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        return reply == QMessageBox.StandardButton.Yes

    def process_files(self):
        """When file confirm button is pressed: double check validity, then process files"""

        # Check all file fields filled out
        all_valid = (
            all(
                section["paradigm_combo"].currentIndex() != 0
                and section["mff_label"].text() != "No file selected"
                for section in self.file_upload_tab.sections
            )
            and self.file_upload_tab.notes_label != "No file selected"
        )

        if not all_valid:
            QMessageBox.warning(
                self,
                "WARNING",
                "Check that the fields are valid. If you see this something is really wrong.",
            )
            return

        if not self.ask_user_for_file_confirmation():
            return

        # initialize progress dialog
        progress_dialog = ProgressDialog(self)
        progress_dialog.show()

        # update data model with file information
        self.file_upload_tab.update_file_info()
        progress_dialog.update_progress(20)

        # save session and file info to deid log
        self.data_model.save_session_to_deid_log()
        progress_dialog.update_progress(40)

        # copy corrected files
        self.data_model.copy_and_rename_files()
        progress_dialog.update_progress(60)

        # copy deid files
        self.data_model.save_deid_files()
        progress_dialog.update_progress(80)

        # zip net placement photos
        self.data_model.save_net_placement_photos()
        progress_dialog.update_progress(100)

        # save sidecar (not used currently)
        # self.data_model.save_sidecar_files()

        # close progress dialog
        progress_dialog.accept()

        # display deid and confirm file transfer
        message = "File transfer complete."
        rand_num = datetime.now().microsecond % 100
        if 5 <= rand_num <= 99: 
            message += r"""File transfer complete. Here is a lucky cat
             /\_/\  
            ( o.o ) 
            > ^ <
            """
        elif 0 <= rand_num <= 4:
            message += r"""File transfer complete. Here is an ultra rare sleepy cat! 
             /\_/\  
            (  -.-  ) zZ
            /       \
            """
        message += f"\n\nYour DeID is: {self.data_model.deid:04}"
        QMessageBox.information(self, "Success", message)

        # reset for next file
        self.reset_app()


class ProgressDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Progress")

        # disable main window when active
        self.setModal(True)

        # disable close button (wait until all transfer is complete)
        self.setWindowFlags(Qt.Window | Qt.WindowTitleHint | Qt.CustomizeWindowHint)

        self.layout = QVBoxLayout(self)

        self.warning_label = QLabel(
            "FILES ARE BEING COPIED, DO NOT TOUCH ANYTHING\n\nUI MAY BECOME UNRESPONSIVE, THAT IS OKAY\n\n",
            self,
        )
        self.layout.addWidget(self.warning_label)

        self.progress_bar = QProgressBar(self)
        self.layout.addWidget(self.progress_bar)

        self.setLayout(self.layout)

    def update_progress(self, value):
        self.progress_bar.setValue(value)
        QApplication.processEvents()


class DataModel:

    PARADIGM_TO_DEID_COLUMN_NAME = {
        "rest": "Resting",
        "resteyesclosed": "Resting",
        "resteyesclosedeyesopen": "Resting",
        "chirp": "Chirp",
        "chirplong": "Chirp",
        "ssct": "Steady State",
        "assr": "ASSR",
        "rleeg": "Reversal Learning",
        "talk": "TalkListen",
        "listen": "TalkListen",
        "vdaudio": "Visual Discrimination",
        "vdnoaudio": "Visual Discrimination",
        "slstructured": "SL Structured",
        "slrandom": "SL Random",
        "habituation": "Habituation",
        "bblong": "BB Long",
        "tactilechirp": "Tactile Chirp",
        "tactilehab": "Tactile Habituation",
        "oddball": "Oddball",
        "other": "Other",
    }

    def __init__(self):

        # Load file path configuration
        self.filepath_dict = self.load_file_paths()
        self.deid_log_filepath = self.filepath_dict["deid_log_filepath"]

        # Load UI configuration containing presets
        self.config_file_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), "ui_config.json"
        )

        # DO NOT CHECK FOR LOCAL BACK UP DISCREPANCIES
        # self.check_if_local_backup_matches_synced_log()

        # Load UI configuration file
        with open(self.config_file_path, "r") as f:
            self.config_dict = json.load(f)

        # Notes file path
        self.notes_file = None

        # Net placement photos path
        self.net_placement_photos = None

        # Session information
        self.session_info = {
            "study": None,
            "visit_number": None,
            "subject_id": None,
            "subject_initials": None,
            "date": None,
            "location": None,
            "net_serial_number": None,
            "cap_type": None,
            "other_notes": None,
        }

        # List of dictionaries containing EEG file paradigm and file paths
        self.eeg_file_info = []

        # Init deid log
        self.deid_log = pd.DataFrame()
        self.load_deid_log(self.deid_log_filepath)

        # DeID for current session
        self.deid = None

    def load_file_paths(self):
        filepath_config_file_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), "filepath_config.json"
        )
        try:
            with open(filepath_config_file_path, "r") as file:
                config = json.load(file)

            for key, path in config.items():
                expanded_path = os.path.expanduser(path)

                # check if path exists
                if not os.path.exists(expanded_path):
                    QMessageBox.critical(
                        None,
                        "ERROR",
                        f"Make sure that the following are connected:\n1. USB\n2. External drive\n3. OneDrive\n\nThe following path cannot be found:\n{expanded_path}",
                    )
                    sys.exit(1)

            expanded_paths = {
                key: os.path.expanduser(path) for key, path in config.items()
            }
            return expanded_paths

        except FileNotFoundError:
            QMessageBox.critical(
                None, "Error", f"JSON file not found: {filepath_config_file_path}"
            )
            sys.exit(1)

    def clear_data(self):
        """Reset data model"""
        self.__init__()

    def get_list_of_current_paradigms(self):
        """Get list of paradigms for current study preset"""
        current_study = self.session_info["study"]
        return self.config_dict[current_study]["paradigm"]["options"]

    def load_deid_log(self, file_path):
        """Read deid log into pandas dataframe, ignoring rows without available deids"""

        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Deid log {file_path} does not exist!")

        # load deid log into data model
        with open(file_path, "rb") as file:
            self.deid_log = pd.read_excel(file, engine="openpyxl")

        # Filter out rows where first column is NaN (no deid available)
        first_column = self.deid_log.columns[0]
        self.deid_log = self.deid_log[self.deid_log[first_column].notna()]
        self.deid_log.reset_index(drop=True, inplace=True)

    def save_session_to_deid_log(self):
        """Get deid and update deid log with current session information"""

        # get deid log and determine first empty row
        df = self.deid_log.copy()
        empty_row_index = self.get_empty_row_index_from_deid_log()

        # check if there's rows available
        if empty_row_index >= len(df):
            raise Exception("No empty rows available in the CSV")

        # set deid from log
        self.deid = self.get_deid(empty_row_index)

        # Update DataFrame with session info
        cur_session_data = {
            "Study": self.session_info["study"],
            "Subject ID": self.session_info["subject_id"],
            "Visit Num": self.session_info["visit_number"],
            "Visit Date": self.session_info["date"],
            "Initials": self.session_info["subject_initials"],
            "Location": self.session_info["location"],
            "Net Serial Number": int(self.session_info["net_serial_number"]),
            "Notes": self.session_info["other_notes"],
        }
        for key, value in cur_session_data.items():
            df.at[empty_row_index, key] = value

        # Add paradigms
        file_names_list = []
        for eeg_file_dict in self.eeg_file_info:
            cur_paradigm = eeg_file_dict["paradigm"]
            column_name = self.PARADIGM_TO_DEID_COLUMN_NAME.get(cur_paradigm, None)
            if column_name:
                if pd.isna(df.at[empty_row_index, column_name]):
                    df.at[empty_row_index, column_name] = 1
                else:
                    df.at[empty_row_index, column_name] += 1
            if "mff_file" in eeg_file_dict:
                file_names_list.append(os.path.basename(eeg_file_dict["mff_file"]))

        # Join collected file names into a semicolon-separated string
        df.at[empty_row_index, "original_file_names"] = ";".join(file_names_list)

        # Update work book (only at specified row)
        wb = load_workbook(self.deid_log_filepath)
        sheet = wb.active

        sheet.protection.disable()

        for col_idx, col_name in enumerate(df.columns, start=1):
            if col_idx == 1:
                continue
            cell = sheet.cell(row=empty_row_index + 2, column=col_idx)  # get cell
            cell.value = df.at[empty_row_index, col_name]  # set cell value
            if col_name == "Visit Date":
                cell.number_format = "MM/DD/YYYY"

        sheet.protection.enable()

        # Save the workbook with the updated row
        wb.save(self.deid_log_filepath)

        # Save a second copy of the workbook as a backup
        wb.save(self.filepath_dict["deid_log_local_backup_filepath"])

        wb.close()

    def check_if_session_info_already_exists(self):
        """Check if a row with the same session data already exists in the DataFrame"""

        # Vectorized comparison for better performance
        mask = (
            (self.deid_log["Study"] == self.session_info["study"])
            & (
                self.deid_log["Subject ID"].astype(str)
                == str(self.session_info["subject_id"])
            )
            & (self.deid_log["Visit Num"] == self.session_info["visit_number"])
        )

        return mask.any()

    def check_if_local_backup_matches_synced_log(self):
        """Check if local copy matches synced copy to ensure there are no conflicts"""

        df1 = pd.read_excel(self.filepath_dict["deid_log_filepath"])[
            ["Study", "Subject ID", "Visit Num"]
        ]
        df2 = pd.read_excel(self.filepath_dict["deid_log_local_backup_filepath"])[
            ["Study", "Subject ID", "Visit Num"]
        ]

        if not df1.equals(df2):
            QMessageBox.critical(
                None,
                "ERROR",
                f"OneDrive Sync Error! Local copy does not match synced deid log. PANIC!!!",
            )
            sys.exit(1)

    def get_empty_row_index_from_deid_log(self):
        """Find the index of the first completely empty row (ignoring the first column)"""
        empty_rows = self.deid_log.loc[:, self.deid_log.columns[1:]].isna().all(axis=1)
        if not empty_rows.any():
            raise ValueError("No available rows in deid log, run out of deids.")
        return empty_rows.idxmax()

    def get_deid(self, row):
        """Get deid from deid log, given a row index"""
        return self.deid_log.at[row, self.deid_log.columns[0]]

    def generate_base_name(self, paradigm, audio_source="", counter=""):
        """Generate base file name with all necessary data."""
        dat = self.session_info
        base_name = f"{dat['study']}_{dat['visit_number']}_{paradigm}{counter}_{dat['subject_id']}_{dat['subject_initials']}_{dat['date']}"

        # Add additional modifiers if needed
        if self.session_info.get("cap_type") == "babycap":
            base_name += "_babycap"
        if audio_source == "speakers":
            base_name += "_speakers"

        return base_name

    def check_file_exists(self, path):
        """Check if a file exists and raise an error if it does."""
        if os.path.exists(path):
            QMessageBox.critical(
                None,
                "ERROR",
                f"File '{path}' already exists. Check that you entered the session info correctly!",
            )
            raise FileExistsError(f"File '{path}' already exists.")

    def copy_and_rename_files(self):
        paradigm_counter = {}
        destination_folder = self.filepath_dict["mff_backup_dir"]
        dat = self.session_info

        for cur_file_info in self.eeg_file_info:
            src_path = cur_file_info["mff_file"]

            # Skip files if paths are missing
            if not src_path:
                continue

            paradigm = cur_file_info["paradigm"]

            # update counter for paradigm
            counter = paradigm_counter.get(paradigm, 0) + 1
            paradigm_counter[paradigm] = counter

            # gemerate base name
            counter_str = "" if counter == 1 else str(counter)
            base_name = self.generate_base_name(paradigm, cur_file_info['audio_source'], counter_str)

            final_directory_path = os.path.join(
                destination_folder,
                dat["study"],
                f"{dat['subject_id']} {dat['subject_initials']}",
                dat["visit_number"],
            )
            os.makedirs(final_directory_path, exist_ok=True)

            # Create destination paths
            dst_path = os.path.join(final_directory_path, base_name + ".mff")

            # Check if files already exist
            self.check_file_exists(dst_path)

            # Copy files
            shutil.copytree(src_path, dst_path)

        # Save notes file
        new_notes_file_name = (
            f"{dat['study']}_{dat['visit_number']}_{dat['subject_id']}_{dat['subject_initials']}_{dat['date']}"
            + os.path.splitext(self.notes_file)[1]
        )
        shutil.copy2(
            self.notes_file, os.path.join(final_directory_path, new_notes_file_name)
        )

    def save_deid_files(self):
        destination_folder = self.filepath_dict["mff_deid_dir"]
        paradigm_counter = {}

        for cur_file_info in self.eeg_file_info:
            src_path = cur_file_info["mff_file"]
            audio_source = cur_file_info['audio_source']

            # Skip files if paths are missing
            if not src_path:
                continue

            paradigm = cur_file_info["paradigm"]

            # update counter for paradigm
            counter = paradigm_counter.get(paradigm, 0) + 1
            paradigm_counter[paradigm] = counter

            # generate base name
            counter_str = "" if counter == 1 else str(counter)
            base_name = f"{self.deid:04}_{paradigm}{counter_str}"

            if self.session_info.get("cap_type") == "babycap":
                base_name += "_babycap"
            if audio_source == "speakers":
                base_name += "_speakers"

            dst_path_deid = os.path.join(
                destination_folder, paradigm, base_name + ".mff"
            )

            # check that you're not overwriting any files
            self.check_file_exists(dst_path_deid)

            # copy deidentified files
            shutil.copytree(src_path, dst_path_deid)

            # deidentify mff files (remove video and original file name)

            # try:
            #     self.deidentify_mff(
            #         mff_file_path = dst_path_deid,
            #         original_filename = os.path.splitext(os.path.basename(src_path))[0],
            #         new_filename = base_name
            #     )
            # except Exception as error:
            #     print("PANIC")

        # Save notes file
        new_notes_file_name = (
            f"{self.deid:04}_notes" + os.path.splitext(self.notes_file)[1]
        )
        shutil.copy2(
            self.notes_file, os.path.join(destination_folder, new_notes_file_name)
        )

    def save_net_placement_photos(self):
        if not self.net_placement_photos:
            return

        destination_folder = self.filepath_dict["net_placement_photo_dir"]
        dat = self.session_info
        paradigm = "netplacementphotos"

        base_name = self.generate_base_name(paradigm)
        dst_path_zip = os.path.join(destination_folder, base_name + ".zip")

        self.check_file_exists(dst_path_zip)

        try:
            with ZipFile(dst_path_zip, "w") as zip_file:
                for image in self.net_placement_photos:
                    zip_file.write(image, os.path.basename(image))
        except Exception as e:
            QMessageBox.critical(
                None, "ERROR", f"Error zipping net placement photos:\n{str(e)}"
            )
            sys.exit(1)

    def deidentify_mff(mff_file_path, original_filename, new_filename):

        # List of files to deidentify within the .MFF directory
        files_to_deidentify = [
            "hostTimes.xml",
            "movieSyncs1.xml",
            "subject.xml",
            "techNote.rtf",
        ]
        # Loop through each file and apply deidentification
        for file_name in files_to_deidentify:
            file_path = os.path.join(mff_file_path, file_name)

            # Check if file exists
            if not os.path.exists(file_path):
                print(f"File not found: {file_path}")
                return

            # Read the file as text
            with open(file_path, "r", encoding="utf-8") as file:
                file_content = file.read()
            # Replace old file name with the deidentified file name
            file_content = file_content.replace(original_filename, new_filename)
            # Replace old ID with deidentified ID
            original_id = original_filename.rsplit("_", 2)[0]
            file_content = file_content.replace(original_id, new_filename)
            # Write the deidentified content back to the file
            with open(file_path, "w", encoding="utf-8") as file:
                file.write(file_content)

        # Loop through files
        for cur_file in os.listdir(mff_file_path):

            # Remove participant video .mov files
            if cur_file.endswith(".mov"):
                movie_file_path = os.path.join(mff_file_path, cur_file)
                os.remove(movie_file_path)
            # Rename log file
            if original_filename in cur_file and cur_file.endswith(".txt"):
                os.rename(
                    os.path.join(mff_file_path, cur_file),
                    os.path.join(
                        mff_file_path, cur_file.replace(original_filename, new_filename)
                    ),
                )

    #################################################
    ############## SAVING SIDECAR JSON ##############
    #################################################

    def save_sidecar_files(self):
        """Save session and file info in json sidecar file"""

        # make copy of session info
        dat = self.session_info.copy()
        sidecar_dict = self.session_info.copy()
        sidecar_dict = {"session_" + key: value for key, value in sidecar_dict.items()}

        # generate ULID
        new_ulid = ulid.new()
        sidecar_dict["session_parent_ulid"] = str(new_ulid)
        sidecar_dict["session_parent_timestamp"] = (
            new_ulid.timestamp().datetime.isoformat()
        )

        # initialize list for file info
        sidecar_dict["session_eeg_file_info"] = []

        # add file info
        for cur_file_info in self.eeg_file_info:

            # add file info to json
            new_file_info = {}
            new_file_info["file_paradigm"] = cur_file_info["paradigm"]
            new_file_info["file_filename"] = cur_file_info["mff_file"]

            # placeholders
            new_file_info["file_azure_json"] = ""
            new_file_info["file_azure_link"] = ""
            new_file_info["file_stage"] = "original"

            sidecar_dict["session_eeg_file_info"].append(new_file_info)

        # Sub directory path for saving files in correct folder
        destination_folder = self.filepath_dict["mff_backup_dir"]
        final_directory_path = os.path.join(
            destination_folder,
            dat["study"],
            dat["subject_id"] + " " + dat["subject_initials"],
            dat["visit_number"],
        )
        os.makedirs(final_directory_path, exist_ok=True)
        base_name = f"{dat['study']}_{dat['visit_number']}_{dat['subject_id']}_{dat['subject_initials']}_{dat['date']}_sessionSidecar"
        dst_path_sidecar = os.path.join(final_directory_path, base_name + ".json")
        with open(dst_path_sidecar, "w") as outfile:
            json.dump(sidecar_dict, outfile, indent=4)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    main_window = MainWindow()
    main_window.show()
    sys.exit(app.exec_())
